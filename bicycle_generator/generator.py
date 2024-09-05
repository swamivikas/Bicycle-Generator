import itertools
import openpyxl
import json

def generate_bicycle_json(file_path):
    wb = openpyxl.load_workbook(file_path)
    
    id_sheet = wb["ID"]
    designators = {}
    
    for col_idx in range(1, id_sheet.max_column + 1):
        header = id_sheet.cell(row=1, column=col_idx).value  # Column header in the first row
        designators[header] = [id_sheet.cell(row=row_idx, column=col_idx).value for row_idx in range(2, id_sheet.max_row + 1) if id_sheet.cell(row=row_idx, column=col_idx).value is not None]

    def generate_combinations(designators):
        for combination in itertools.product(*designators.values()):
            yield combination

    general_sheet = wb["GENERAL"]
    general_fields = {general_sheet.cell(row=1, column=i).value: general_sheet.cell(row=2, column=i).value for i in range(1, general_sheet.max_column + 1)}
    
    output = []
    
    for combination in generate_combinations(designators):
        bicycle = {}
        bicycle['ID'] = '-'.join(map(str, combination))  
        bicycle.update(general_fields)
        
        for sheet_name in wb.sheetnames[2:]:  
            sheet = wb[sheet_name]
            designator_column = [sheet.cell(row=i, column=1).value for i in range(1, sheet.max_row + 1)]
            
            for i, value in enumerate(combination):
                if value in designator_column:
                    row_index = designator_column.index(value) + 1
                    for col_index in range(2, sheet.max_column + 1):
                        field_name = sheet.cell(row=1, column=col_index).value
                        field_value = sheet.cell(row=row_index, column=col_index).value
                        bicycle[field_name] = field_value
        
        output.append(bicycle)

    return json.dumps(output, indent=4)
